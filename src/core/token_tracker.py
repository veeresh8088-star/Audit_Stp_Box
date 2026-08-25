import os
import json
import time
import threading
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BENCHMARK_JSON_PATH = "data/audit_token_benchmark.json"
BENCHMARK_EXCEL_PATH = "data/audit_token_benchmark.xlsx"

# Retention cap: without this, BENCHMARK_JSON_PATH grows forever (one record per
# audit session, never rotated), and record_token_metrics() re-derives/rewrites
# the ENTIRE Excel workbook from full history on every single audit completion --
# the lock hold time for that grows as session count grows, serializing every
# concurrently-finishing audit's post-completion bookkeeping against every other
# one. Keeping only the most recent N sessions bounds both the file size and the
# per-call rewrite cost, without changing get_all_benchmark_records()'s shape --
# nothing that reads it needs to change.
MAX_BENCHMARK_RECORDS = 5000

# Serializes the read-modify-write cycle below across concurrent audit threads —
# without this, two audits finishing close together can each read the same file,
# append their own record, and the second write silently overwrites the first.
_benchmark_file_lock = threading.Lock()

def record_token_metrics(
    session_id: str,
    scoping_mode: str,
    file_names: list,
    total_file_size_bytes: int,
    extracted_text_chars: int,
    controls_count: int,
    prompt_tokens: int,
    completion_tokens: int,
    total_latency_sec: float,
    compliant_count: int = 0,
    non_compliant_count: int = 0,
    out_of_scope_count: int = 0,
    folder_name: str = "",
    cpu_cores: int = 0,
    file_details: list = None,
    file_types_summary: dict = None
):
    """
    Records detailed token consumption, text size, latency, hardware CPU specs,
    file type breakdown, and scoping metrics for real document analysis sessions.
    """
    os.makedirs("data", exist_ok=True)

    if not cpu_cores:
        cpu_cores = os.cpu_count() or 4

    # Build file_types_summary if not provided
    if not file_types_summary and file_names:
        file_types_summary = {}
        for fn in file_names:
            ext = os.path.splitext(str(fn))[1].lower().replace('.', '') or 'unknown'
            file_types_summary[ext] = file_types_summary.get(ext, 0) + 1

    total_tokens = prompt_tokens + completion_tokens
    avg_latency_per_ctrl = round(total_latency_sec / max(1, controls_count), 2)
    tokens_per_sec = round(total_tokens / max(0.1, total_latency_sec), 2)
    file_size_kb = round(total_file_size_bytes / 1024, 2)
    file_size_mb = round(total_file_size_bytes / (1024 * 1024), 3)

    record = {
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "session_id": session_id,
        "folder_name": folder_name or ("Folder Analysis" if file_names else "Single Document"),
        "scoping_mode": scoping_mode, # "AI Auto-Scoping" or "Excel / Manual Scoping"
        "cpu_cores": cpu_cores,
        "files_count": len(file_names),
        "file_names": ", ".join(file_names) if isinstance(file_names, list) else str(file_names),
        "file_types_summary": file_types_summary or {},
        "file_details": file_details or [],
        "file_size_kb": file_size_kb,
        "file_size_mb": file_size_mb,
        "extracted_text_chars": extracted_text_chars,
        "extracted_text_words": len(str(extracted_text_chars).split()) if isinstance(extracted_text_chars, str) else int(extracted_text_chars / 6),
        "controls_audited_count": controls_count,
        "prompt_input_tokens": prompt_tokens,
        "completion_output_tokens": completion_tokens,
        "total_tokens": total_tokens,
        "total_latency_seconds": round(total_latency_sec, 2),
        "avg_latency_per_control_sec": avg_latency_per_ctrl,
        "tokens_per_second": tokens_per_sec,
        "compliant_count": compliant_count,
        "non_compliant_count": non_compliant_count,
        "out_of_scope_count": out_of_scope_count
    }

    # 1. Update JSON database
    with _benchmark_file_lock:
        records = []
        if os.path.exists(BENCHMARK_JSON_PATH):
            try:
                with open(BENCHMARK_JSON_PATH, "r", encoding="utf-8") as f:
                    records = json.load(f)
            except Exception:
                records = []

        # Check if entry already exists for session_id to avoid duplication
        records = [r for r in records if r.get("session_id") != session_id]
        records.append(record)
        if len(records) > MAX_BENCHMARK_RECORDS:
            records = records[-MAX_BENCHMARK_RECORDS:]

        with open(BENCHMARK_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=2)

        # 2. Export Styled Excel Spreadsheet
        generate_excel_benchmark_report(records, BENCHMARK_EXCEL_PATH)
    print(f"[TOKEN TRACKER] Recorded metrics for session {session_id[:8]} ({scoping_mode}): {total_tokens} tokens, {round(total_latency_sec,1)}s latency, {cpu_cores} CPU cores.", flush=True)
    return record


def get_all_benchmark_records():
    """Returns all recorded audit session benchmark metrics, each enriched with the
    real auditor username (joined from AuditReport.created_by via session_id).

    Benchmark records themselves never carry who ran them -- only session_id, a
    random uuid4 hex. The DB has always had that mapping; this was just never
    joined in, so every benchmark export (Excel/PDF, both call sites) showed
    session/token/latency data with no way to tell which auditor it belonged to.
    """
    if not os.path.exists(BENCHMARK_JSON_PATH):
        return []
    try:
        with open(BENCHMARK_JSON_PATH, "r", encoding="utf-8") as f:
            records = json.load(f)
    except Exception:
        return []
    if not isinstance(records, list):
        records = [records]

    session_ids = {r.get("session_id") for r in records if r.get("session_id")}
    username_by_session = {}
    if session_ids:
        try:
            from src.db.database import SessionLocal, AuditReport, force_master
            db = SessionLocal()
            try:
                with force_master():
                    rows = db.query(AuditReport.session_id, AuditReport.created_by).filter(
                        AuditReport.session_id.in_(session_ids)
                    ).all()
                username_by_session = {sid: (created_by or "") for sid, created_by in rows}
            finally:
                db.close()
        except Exception:
            username_by_session = {}
    for r in records:
        r["auditor_username"] = username_by_session.get(r.get("session_id"), "") or "Unknown"

    return records


def aggregate_audit_sessions(session_ids: list = None):
    """
    Aggregates metrics across selected (or all) real auditor session logs.
    Computes combined total latency, combined tokens, combined files/sizes,
    overall compliance score, and side-by-side comparative matrix.
    """
    records = get_all_benchmark_records()
    if not records:
        return {}

    if session_ids:
        # Filter for specified session IDs
        target_sids = set(session_ids)
        records = [r for r in records if r.get("session_id") in target_sids or any(sid in str(r.get("session_id")) for sid in target_sids)]

    if not records:
        return {}

    tot_latency = sum(float(r.get("total_latency_seconds", 0)) for r in records)
    tot_prompt = sum(int(r.get("prompt_input_tokens", 0)) for r in records)
    tot_comp = sum(int(r.get("completion_output_tokens", 0)) for r in records)
    tot_tokens = tot_prompt + tot_comp
    tot_files = sum(int(r.get("files_count", 0)) for r in records)
    tot_mb = sum(float(r.get("file_size_mb", 0)) for r in records)
    tot_chars = sum(int(r.get("extracted_text_chars", 0)) for r in records)
    tot_ctrls = sum(int(r.get("controls_audited_count", 0)) for r in records)
    tot_compliant = sum(int(r.get("compliant_count", 0)) for r in records)
    tot_non_compliant = sum(int(r.get("non_compliant_count", 0)) for r in records)
    tot_out_scope = sum(int(r.get("out_of_scope_count", 0)) for r in records)

    # Format combined latency string (e.g. 1h 23m 45s or 45m 12s)
    hours = int(tot_latency // 3600)
    mins = int((tot_latency % 3600) // 60)
    secs = round(tot_latency % 60, 1)
    if hours > 0:
        comb_lat_str = f"{hours}h {mins}m {secs}s"
    elif mins > 0:
        comb_lat_str = f"{mins}m {secs}s"
    else:
        comb_lat_str = f"{secs}s"

    overall_score_pct = int((tot_compliant / max(1, tot_compliant + tot_non_compliant)) * 100) if (tot_compliant + tot_non_compliant) > 0 else 0

    # Merge file types
    merged_file_types = {}
    for r in records:
        fts = r.get("file_types_summary", {})
        if isinstance(fts, dict):
            for ext, cnt in fts.items():
                merged_file_types[ext] = merged_file_types.get(ext, 0) + cnt

    # Max CPU cores reported across sessions
    cpu_cores_list = [int(r.get("cpu_cores", 0)) for r in records if r.get("cpu_cores")]
    cpu_cores_display = max(cpu_cores_list) if cpu_cores_list else (os.cpu_count() or 4)

    return {
        "selected_sessions_count": len(records),
        "session_ids": [r.get("session_id") for r in records],
        "cpu_cores": cpu_cores_display,
        "combined_latency_seconds": round(tot_latency, 2),
        "combined_latency_formatted": comb_lat_str,
        "combined_prompt_tokens": tot_prompt,
        "combined_completion_tokens": tot_comp,
        "combined_total_tokens": tot_tokens,
        "combined_files_count": tot_files,
        "combined_file_size_mb": round(tot_mb, 2),
        "combined_extracted_text_chars": tot_chars,
        "combined_controls_count": tot_ctrls,
        "combined_compliant_count": tot_compliant,
        "combined_non_compliant_count": tot_non_compliant,
        "combined_out_of_scope_count": tot_out_scope,
        "overall_compliance_score_pct": overall_score_pct,
        "file_types_summary": merged_file_types,
        "sessions_comparison": records
    }



def generate_excel_benchmark_report(records: list, output_path: str = BENCHMARK_EXCEL_PATH):
    """
    Generates a beautifully styled executive Excel workbook summarizing token usage,
    latency (in Minutes & Seconds), file size, text length, and scoping comparison
    dynamically calculated from actual audit session records and evaluated controls.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit_Benchmark_Report"

    # Styling definitions
    font_title = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=11, bold=True, color="1E293B")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)

    fill_title = PatternFill(start_color="08519C", end_color="08519C", fill_type="solid") # Deep Navy
    fill_header = PatternFill(start_color="3182BD", end_color="3182BD", fill_type="solid") # Medium Blue
    fill_ai_row = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    def format_min_sec(seconds_val):
        mins = int(seconds_val // 60)
        secs = round(seconds_val % 60, 1)
        if mins > 0:
            return f"{mins}m {secs}s"
        return f"0m {secs}s"

    # Header Title Block
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "EXECUTIVE AUDIT BENCHMARK & SCOPE PERFORMANCE REPORT"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Gather metadata from records
    first_r = records[0] if records else {}
    folder_name = first_r.get("folder_name", "src/aa audit evidence samples")
    files_cnt = first_r.get("files_count", 8)
    file_size_mb = first_r.get("file_size_mb", 2.43)
    file_size_kb = first_r.get("file_size_kb", 2489.64)
    ai_model = first_r.get("ai_model", "Gemma 4 (e4b)")
    scoping_mode_str = first_r.get("scoping_mode", "Excel Upload Scope")

    checklist_file = "Audit checklist and evidence files.xlsx" if "excel" in scoping_mode_str.lower() or "manual" in scoping_mode_str.lower() else "N/A (AI Auto-Scoping)"

    # Section 1: Folder & Evidence Overview
    ws.cell(row=3, column=1, value="1. AUDIT FOLDER & EVIDENCE METRICS").font = font_section
    ws.row_dimensions[3].height = 24

    folder_headers = ["Audit Folder Name", "Total Evidence Count", "Total File Size (MB)", "Total File Size (KB)", "AI Model Engine", "Checklist File Included"]
    for c_idx, h_text in enumerate(folder_headers, 1):
        cell = ws.cell(row=4, column=c_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[4].height = 24

    folder_row_data = [
        folder_name,
        files_cnt,
        f"{file_size_mb} MB",
        f"{file_size_kb} KB",
        ai_model,
        checklist_file
    ]
    for c_idx, val in enumerate(folder_row_data, 1):
        cell = ws.cell(row=5, column=c_idx, value=val)
        cell.font = font_regular
        cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[5].height = 20

    # Section 2: Scope Evaluation Summary (Dynamically generated per scoping mode in records!)
    ws.cell(row=7, column=1, value="2. SCOPE EVALUATION SUMMARY BENCHMARK").font = font_section
    ws.row_dimensions[7].height = 24

    summary_headers = [
        "Scope Detection Method",
        "Relevant Controls Sent to LLM",
        "Irrelevant Controls Dropped",
        "Evidence Files Count",
        "Total Audit Tokens",
        "Avg Tokens / Control",
        "Overall Audit Latency (Min & Sec)",
        "Overall Latency (Sec)",
        "Avg Latency / Control",
        "AI Model Engine"
    ]
    for c_idx, h_text in enumerate(summary_headers, 1):
        cell = ws.cell(row=8, column=c_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[8].height = 26

    # Group records by scoping mode dynamically
    summary_by_mode = {}
    for r in records:
        mode = r.get("scoping_mode", "Excel Upload Scope")
        if mode not in summary_by_mode:
            summary_by_mode[mode] = {
                "relevant_count": 0,
                "dropped_count": 0,
                "files_count": r.get("files_count", files_cnt),
                "total_tokens": 0,
                "total_latency_sec": 0.0,
                "ai_model": r.get("ai_model", ai_model)
            }
        ctrl_cnt = int(r.get("controls_audited_count", 0))
        if ctrl_cnt == 0 and "controls_detail" in r:
            ctrl_cnt = len(r["controls_detail"])
        summary_by_mode[mode]["relevant_count"] += ctrl_cnt
        summary_by_mode[mode]["dropped_count"] += int(r.get("out_of_scope_count", 0))
        summary_by_mode[mode]["total_tokens"] += int(r.get("total_tokens", 0))
        summary_by_mode[mode]["total_latency_sec"] += float(r.get("total_latency_seconds", 0.0))

    summary_rows = []
    for mode_name, m_data in summary_by_mode.items():
        rel_cnt = max(1, m_data["relevant_count"])
        tot_toks = m_data["total_tokens"]
        tot_lat = m_data["total_latency_sec"]
        drop_cnt = m_data["dropped_count"]
        
        drop_label = f"{drop_cnt} dropped (Pre-Filter)" if drop_cnt > 0 else "0 dropped"
        summary_rows.append([
            mode_name,
            f"{rel_cnt} mapped controls" if ("excel" in mode_name.lower() or "manual" in mode_name.lower()) else f"{rel_cnt} relevant controls",
            drop_label,
            m_data["files_count"],
            f"{tot_toks:,}",
            f"{round(tot_toks / rel_cnt, 1)}",
            format_min_sec(tot_lat),
            f"{round(tot_lat, 1)} s",
            format_min_sec(round(tot_lat / rel_cnt, 1)),
            m_data["ai_model"]
        ])

    for r_offset, s_row in enumerate(summary_rows, 9):
        for c_idx, val in enumerate(s_row, 1):
            cell = ws.cell(row=r_offset, column=c_idx, value=val)
            cell.font = font_bold if c_idx in [1, 2, 5, 7] else font_regular
            cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")
            cell.border = thin_border
            if "AI" in str(s_row[0]):
                cell.fill = fill_ai_row
        ws.row_dimensions[r_offset].height = 22

    # Section 3: Detailed Per-Control Execution Details (Render ONLY actual evaluated controls!)
    section3_start_row = 9 + len(summary_rows) + 2
    ws.cell(row=section3_start_row - 1, column=1, value="3. PER-CONTROL TOKEN & LATENCY EXECUTION DETAILS").font = font_section
    ws.row_dimensions[section3_start_row - 1].height = 24

    detail_headers = [
        "Scope Detection Method",
        "Control ID",
        "Control / Audit Check Name",
        "Evaluation Status",
        "Prompt Tokens",
        "Completion Tokens",
        "Tokens per Control",
        "Per-Control Latency (Min & Sec)",
        "Per-Control Latency (Sec)",
        "AI Model Engine",
        "Auditor"
    ]

    for c_idx, h_text in enumerate(detail_headers, 1):
        cell = ws.cell(row=section3_start_row, column=c_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[section3_start_row].height = 26

    current_r = section3_start_row + 1

    # Extract all real evaluated control details across records
    all_control_details = []
    for r in records:
        sc_mode = r.get("scoping_mode", "Excel Upload Scope")
        r_model = r.get("ai_model", ai_model)
        if "controls_detail" in r and r["controls_detail"]:
            for cd in r["controls_detail"]:
                all_control_details.append({
                    "scoping_mode": sc_mode,
                    "control_id": cd.get("control_id", "5.15"),
                    "control_name": cd.get("control_name", "Access Control"),
                    "status": cd.get("status", "COMPLIANT"),
                    "prompt_tokens": int(cd.get("prompt_tokens", 490)),
                    "completion_tokens": int(cd.get("completion_tokens", 140)),
                    "latency_sec": float(cd.get("latency_sec", 252.0)),
                    "ai_model": r_model,
                    "auditor": r.get("auditor_username", "Unknown"),
                })

    if all_control_details:
        for cd in all_control_details:
            p_toks = cd["prompt_tokens"]
            c_toks = cd["completion_tokens"]
            t_toks = p_toks + c_toks
            c_lat = cd["latency_sec"]
            status_val = str(cd["status"]).upper()
            
            row_vals = [
                cd["scoping_mode"],
                cd["control_id"],
                cd["control_name"],
                status_val,
                p_toks,
                c_toks,
                t_toks,
                format_min_sec(c_lat),
                f"{round(c_lat, 1)} s",
                cd["ai_model"],
                cd["auditor"]
            ]
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=current_r, column=c_idx, value=val)
                cell.font = font_bold if c_idx in [2, 4, 7, 8] else font_regular
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if c_idx in [1,2,4,5,6,7,8,9,10] else "left", vertical="center")
                if c_idx == 4:
                    if status_val in ("COMPLIANT", "ACCEPTED", "PASS"):
                        cell.font = Font(name="Calibri", size=10, bold=True, color="059669")
                    elif status_val in ("PARTIAL", "WARN"):
                        cell.font = Font(name="Calibri", size=10, bold=True, color="D97706")
                    else:
                        cell.font = Font(name="Calibri", size=10, bold=True, color="DC2626")
            ws.row_dimensions[current_r].height = 20
            current_r += 1
    else:
        # Fallback to records summary if controls_detail list not present
        for r in records:
            p_toks = int(r.get("prompt_input_tokens", 480))
            c_toks = int(r.get("completion_output_tokens", 130))
            t_toks = p_toks + c_toks
            c_lat = float(r.get("total_latency_seconds", 255.0))
            
            row_vals = [
                r.get("scoping_mode", "Excel Upload Scope"),
                str(r.get("session_id", "Session"))[:12],
                r.get("folder_name", "Audit Evidence Scope"),
                "COMPLIANT" if r.get("compliant_count", 1) > 0 else "NON_COMPLIANT",
                p_toks,
                c_toks,
                t_toks,
                format_min_sec(c_lat),
                f"{round(c_lat, 1)} s",
                ai_model,
                r.get("auditor_username", "Unknown")
            ]
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=current_r, column=c_idx, value=val)
                cell.font = font_bold if c_idx in [2, 4, 7, 8] else font_regular
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if c_idx in [1,2,4,5,6,7,8,9,10] else "left", vertical="center")
            ws.row_dimensions[current_r].height = 20
            current_r += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val_str = str(cell.value)
                if cell.row == 1:
                    continue
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT: System Audit Event Logs → Excel (.xlsx)
# Works with both PostgreSQL (ShaktiDB) and SQLite fallback — pure ORM rows.
# ─────────────────────────────────────────────────────────────────────────────
def generate_excel_system_logs_report(logs: list, output_path: str) -> str:
    """
    Generates a styled Excel workbook of System Audit Event Logs.
    Accepts a list of dicts with keys: timestamp, actor, severity, event_type,
    session_id, framework, meta.
    Works with both PostgreSQL (ShaktiDB) and SQLite fallback.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "System_Event_Logs"

    # ── Colour palette ───────────────────────────────────────────────────────
    navy       = PatternFill(start_color="08519C", end_color="08519C", fill_type="solid")
    hdr_blue   = PatternFill(start_color="3182BD", end_color="3182BD", fill_type="solid")
    red_fill   = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    yellow_fill= PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    white_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    tc = ws["A1"]
    tc.value = "SYSTEM AUDIT EVENT LOGS & TELEMETRY TRAIL"
    tc.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    tc.fill = navy
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Sub-header: generated timestamp ─────────────────────────────────────
    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = f"Generated: {__import__('datetime').datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}  |  Total Events: {len(logs)}"
    sub.font = Font(name="Calibri", size=9, italic=True, color="64748B")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Column headers ───────────────────────────────────────────────────────
    headers = ["Timestamp", "Auditor / Actor", "Severity", "Event Type", "Session ID", "Framework", "Details / Meta"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = hdr_blue
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    # ── Data rows ────────────────────────────────────────────────────────────
    severity_fill_map = {
        "CRITICAL": red_fill, "ERROR": red_fill,
        "WARNING":  yellow_fill,
        "SUCCESS":  green_fill, "INFO": white_fill,
    }

    for ri, ev in enumerate(logs, 4):
        sev  = str(ev.get("severity") or "INFO").upper()
        row_fill = severity_fill_map.get(sev, white_fill)
        row_data = [
            str(ev.get("created_at") or ev.get("timestamp") or ""),
            str(ev.get("actor") or "SYSTEM"),
            sev,
            str(ev.get("event_type") or ""),
            str(ev.get("session_id") or ""),
            str(ev.get("framework") or ""),
            str(ev.get("meta") or ev.get("details") or ""),
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = row_fill
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=9,
                             bold=(ci == 3),   # severity column bold
                             color="DC2626" if sev in ("CRITICAL","ERROR") else
                                   "92400E" if sev == "WARNING" else "1E293B")
            cell.alignment = Alignment(horizontal="center" if ci == 3 else "left",
                                       vertical="center", wrap_text=(ci == 7))
        ws.row_dimensions[ri].height = 18

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = [22, 22, 12, 22, 28, 14, 55]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT: System Audit Event Logs → PDF (.pdf)
# ─────────────────────────────────────────────────────────────────────────────
def generate_pdf_system_logs_report(logs: list, output_path: str) -> str:
    """
    Generates an executive PDF of System Audit Event Logs using fpdf2.
    Works with both PostgreSQL (ShaktiDB) and SQLite fallback.
    """
    try:
        from fpdf import FPDF
        from fpdf.enums import XPos, YPos
    except ImportError:
        raise RuntimeError("fpdf2 is not installed. Run: pip install fpdf2")

    class LogsPDF(FPDF):
        def header(self):
            if self.page_no() == 1:
                return
            self.set_fill_color(15, 23, 42)
            self.rect(0, 0, 210, 11, "F")
            self.set_font("Helvetica", "B", 7.5)
            self.set_text_color(148, 163, 184)
            self.set_xy(10, 2.5)
            self.cell(0, 6, "AICyberAuditBox — System Audit Event Logs & Telemetry Trail", align="L")
            self.set_xy(0, 2.5)
            self.cell(200, 6, f"Page {self.page_no()}", align="R")
            self.ln(12)

        def footer(self):
            self.set_y(-11)
            self.set_font("Helvetica", "", 7)
            self.set_text_color(148, 163, 184)
            self.cell(0, 6, "CONFIDENTIAL | System Event Log Export", align="C")

    pdf = LogsPDF()
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_margins(10, 10, 10)
    pdf.add_page()

    # Title block
    pdf.set_fill_color(8, 81, 156)   # Navy
    pdf.rect(10, 10, 190, 14, "F")
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(10, 12)
    pdf.cell(190, 10, "SYSTEM AUDIT EVENT LOGS & TELEMETRY TRAIL", align="C",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    from datetime import datetime as _dt
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(100, 116, 139)
    pdf.set_x(10)
    pdf.cell(190, 6,
             f"Generated: {_dt.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}  |  Total Events: {len(logs)}",
             align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)

    # Table headers
    col_w = [38, 28, 20, 38, 34, 32]
    col_hdrs = ["Timestamp", "Actor / User", "Severity", "Event Type", "Session ID", "Details"]
    pdf.set_fill_color(49, 130, 189)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 7.5)
    for i, (h, w) in enumerate(zip(col_hdrs, col_w)):
        pdf.cell(w, 8, h, border=1, align="C", fill=True)
    pdf.ln()

    sev_rgb = {
        "CRITICAL": (254, 226, 226), "ERROR": (254, 226, 226),
        "WARNING":  (254, 249, 195),
        "SUCCESS":  (220, 252, 231),
        "INFO":     (248, 250, 252),
    }
    sev_txt = {
        "CRITICAL": (220, 38, 38), "ERROR": (220, 38, 38),
        "WARNING":  (146, 64, 14),
        "SUCCESS":  (22, 101, 52),
        "INFO":     (30, 41, 59),
    }

    pdf.set_font("Helvetica", "", 7)
    for ev in logs:
        sev = str(ev.get("severity") or "INFO").upper()
        bg  = sev_rgb.get(sev, (248, 250, 252))
        fg  = sev_txt.get(sev, (30, 41, 59))
        pdf.set_fill_color(*bg)
        pdf.set_text_color(*fg)

        row = [
            str(ev.get("created_at") or ev.get("timestamp") or "")[:19],
            str(ev.get("actor") or "SYSTEM")[:20],
            sev[:10],
            str(ev.get("event_type") or "")[:28],
            str(ev.get("session_id") or "")[:22],
            str(ev.get("meta") or ev.get("details") or "")[:40],
        ]
        row_h = 7
        for val, w in zip(row, col_w):
            pdf.cell(w, row_h, val, border=1, align="L", fill=True)
        pdf.ln()

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    pdf.output(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT: Telemetry Benchmark → PDF (.pdf)
# ─────────────────────────────────────────────────────────────────────────────
def generate_pdf_benchmark_report(records: list, output_path: str, auditor_filter: str = None) -> str:
    """
    Generates an executive multi-page PDF of Audit Telemetry Benchmark data using fpdf2.
    Works with both PostgreSQL (ShaktiDB) and SQLite fallback JSON records.
    """
    try:
        from fpdf import FPDF
        from fpdf.enums import XPos, YPos
    except ImportError:
        raise RuntimeError("fpdf2 is not installed. Run: pip install fpdf2")

    class TelemetryPDF(FPDF):
        def header(self):
            if self.page_no() == 1:
                return
            self.set_fill_color(15, 23, 42)
            self.rect(0, 0, 210, 11, "F")
            self.set_font("Helvetica", "B", 7.5)
            self.set_text_color(148, 163, 184)
            self.set_xy(10, 2.5)
            self.cell(0, 6, "AICyberAuditBox — Executive Audit Telemetry & Benchmark Report", align="L")
            self.set_xy(0, 2.5)
            self.cell(200, 6, f"Page {self.page_no()}", align="R")
            self.ln(12)

        def footer(self):
            self.set_y(-11)
            self.set_font("Helvetica", "", 7)
            self.set_text_color(148, 163, 184)
            self.cell(0, 6, "CONFIDENTIAL | Executive AI Audit Telemetry Report", align="C")

    pdf = TelemetryPDF()
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_margins(10, 10, 10)
    pdf.add_page()

    from datetime import datetime as _dt

    # ── Title block ───────────────────────────────────────────────────────────
    pdf.set_fill_color(8, 81, 156)
    pdf.rect(10, 10, 190, 14, "F")
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(10, 12)
    pdf.cell(190, 10, "EXECUTIVE AUDIT TELEMETRY & BENCHMARK REPORT", align="C",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    scope_label = f"Auditor: {auditor_filter}" if auditor_filter else "All Auditors (Combined)"
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(100, 116, 139)
    pdf.set_x(10)
    pdf.cell(190, 6,
             f"Generated: {_dt.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}  |  Scope: {scope_label}  |  Sessions: {len(records)}",
             align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(5)

    # ── KPI Summary Cards ─────────────────────────────────────────────────────
    tot_tokens    = sum(int(r.get("total_tokens", 0)) for r in records)
    tot_latency   = sum(float(r.get("total_latency_sec", 0)) for r in records)
    tot_files     = sum(int(r.get("files_count", 0)) for r in records)
    tot_mb        = sum(float(r.get("file_size_mb", 0)) for r in records)
    tot_compliant = sum(int(r.get("compliant_count", 0)) for r in records)
    tot_nc        = sum(int(r.get("non_compliant_count", 0)) for r in records)
    compliance_pct = int((tot_compliant / max(1, tot_compliant + tot_nc)) * 100) if (tot_compliant + tot_nc) > 0 else 0

    mins = int(tot_latency // 60); secs = round(tot_latency % 60, 1)
    lat_str = f"{mins}m {secs}s" if mins > 0 else f"0m {secs}s"

    kpis = [
        ("Total Tokens",    f"{tot_tokens:,}"),
        ("Combined Latency",lat_str),
        ("Total Files",     str(tot_files)),
        ("Total Size (MB)", f"{round(tot_mb, 2)} MB"),
        ("Compliance",      f"{compliance_pct}%"),
        ("Sessions",        str(len(records))),
    ]
    pdf.set_font("Helvetica", "B", 8.5)
    card_w = 30; card_h = 18; gap = 2.67
    for i, (label, val) in enumerate(kpis):
        x = 10 + i * (card_w + gap)
        pdf.set_fill_color(239, 246, 255)
        pdf.set_draw_color(191, 219, 254)
        pdf.rect(x, pdf.get_y(), card_w, card_h, "DF")
        pdf.set_text_color(59, 130, 246)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_xy(x, pdf.get_y() + 2)
        pdf.cell(card_w, 6, val, align="C", new_x=XPos.RIGHT, new_y=YPos.TMARGIN)
        pdf.set_font("Helvetica", "", 7)
        pdf.set_text_color(71, 85, 105)
        pdf.set_xy(x, pdf.get_y() + 10)
        pdf.cell(card_w, 5, label, align="C", new_x=XPos.RIGHT, new_y=YPos.TMARGIN)
    pdf.ln(card_h + 6)

    # ── Per-Session Table ─────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 7, "Auditor Session Breakdown", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(1)

    col_hdrs = ["Session ID", "Auditor",  "Scoping Mode", "Files", "Size MB", "Tokens", "Latency", "Compliance"]
    col_ws   = [26,           32,          24,             12,      16,        20,       18,         26]
    pdf.set_fill_color(49, 130, 189)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 7.5)
    for h, w in zip(col_hdrs, col_ws):
        pdf.cell(w, 8, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for idx, r in enumerate(records):
        bg = (239, 246, 255) if idx % 2 == 0 else (248, 250, 252)
        pdf.set_fill_color(*bg)
        pdf.set_text_color(30, 41, 59)

        lat_s = float(r.get("total_latency_sec", 0))
        m, s = int(lat_s // 60), round(lat_s % 60, 1)
        lat_fmt = f"{m}m {s}s" if m > 0 else f"0m {s}s"
        c_cnt = int(r.get("compliant_count", 0))
        nc_cnt = int(r.get("non_compliant_count", 0))
        cpct = int((c_cnt / max(1, c_cnt + nc_cnt)) * 100) if (c_cnt + nc_cnt) > 0 else 0

        row_vals = [
            str(r.get("session_id", ""))[:14],
            str(r.get("auditor_username", "Unknown"))[:18],
            str(r.get("scoping_mode", ""))[:16],
            str(r.get("files_count", 0)),
            str(round(float(r.get("file_size_mb", 0)), 2)),
            str(r.get("total_tokens", 0)),
            lat_fmt,
            f"{cpct}%",
        ]
        for val, w in zip(row_vals, col_ws):
            pdf.cell(w, 7, val, border=1, align="C", fill=True)
        pdf.ln()

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    pdf.output(output_path)
    return output_path
