# -*- coding: utf-8 -*-
"""
Excel Scoping Parser
====================
Dynamically parses any ISO 27001 audit scoping checklist from an Excel file.

Features:
- Auto-detects 1 to N file columns from the header row (keywords: file, evidence,
  policy, document, attachment, doc, upload)
- Resolves ISO control IDs using 3 fallback steps:
    1. Direct control ID match (e.g. "5.15", "ISO 8.17")
    2. Control name keyword match against USE_CASES
    3. Embedding cosine-similarity against USE_CASES descriptions (no LLM needed)
- Returns a list of dicts: {question, files, control_id, control_label,
  expected_evidence, prompt_hint, severity}
- Works for 8 rows, 50 rows, 100+ rows
"""

import re
import os
from typing import List, Dict, Optional

# ── Keywords that identify file-reference columns in the Excel header ─────────
# IMPORTANT: 'type' columns (File type, Document type) must be EXCLUDED because
# they contain extensions like 'PNG', 'JPG', 'PDF' which are not filenames.
_FILE_COL_KEYWORDS = {
    "file", "evidence", "policy", "document", "attachment", "doc", "upload",
    "exhibit", "reference", "source", "artifact"
}
# Words that, when present, indicate a TYPE/FORMAT column (not a filename column)
_FILE_TYPE_EXCLUSION_KEYWORDS = {"type", "format", "extension", "ext", "kind"}

# ── Direct keyword-to-control mapping for common audit questions ──────────────
# These cover the most frequent audit check phrasings so they resolve instantly
# without needing embedding similarity (which requires Ollama to be running).
_DIRECT_KEYWORD_CONTROL_MAP = [
    # Keywords in question text               -> control use_case string
    (["ntp", "clock", "synchroni", "time server"], "8.17 Clock Synchronization"),
    (["fraud analytics"], "5.1 Policies for Information Security"),
    (["multifactor", "mfa", "multi-factor", "2fa", "two-factor"], "8.5 Secure Authentication"),
    (["pam", "privileged access", "pim", "idam"], "8.2 Privileged Access Rights"),
    (["access control policy", "access control"], "5.15 Access Control"),
    (["authentication", "how is the auth"], "8.5 Secure Authentication"),
    (["asset management policy", "asset management", "asset inventory"], "5.9 Inventory of Information and Other Associated Assets"),
    (["incident management policy", "incident response", "incident plan", "irp"], "5.24 Information Security Incident Management Planning and Preparation"),
    (["business continuity plan", "bcp", "disaster recovery", "dr plan"], "5.29 Information Security During Disruption"),
    (["vulnerability", "patch", "scan"], "8.8 Management of Technical Vulnerabilities"),
    (["hr security policy", "people security policy", "screening"], "6.1 Screening"),
    (["physical security policy", "physical security perimeter", "physical perimeter", "perimeter security"], "7.1 Physical Security Perimeters"),
    (["security monitoring", "log monitoring", "siem", "monitoring activities", "monitoring"], "8.16 Monitoring Activities"),
    (["cpu", "memory", "disk", "utilization", "capacity", "cloudwatch"], "8.6 Capacity Management"),
    (["log archival", "log archived", "archiv", "records retention", "records management"], "5.33 Protection of Records"),
    (["backup", "recovery", "restore"], "8.13 Information Backup"),
    (["incident", "response", "breach"], "5.24 Information Security Incident Management Planning and Preparation"),
    (["encryption", "tls", "ssl", "cipher"], "8.24 Use of Cryptography"),
    (["password", "credential", "secret"], "5.17 Authentication Information"),
    (["firewall", "network security", "network traffic", "firewall policy"], "8.20 Network Security"),
    (["gdpr", "pii", "personal data", "privacy"], "5.34 Privacy and Protection of Personally Identifiable Information (Pii)"),

    # ── Clause 5: Organizational Controls (all previously missing) ────────────
    (["information security policy", "isms policy", "security policy document", "iprotect"], "5.1 Policies for Information Security"),
    (["roles and responsibilities", "isms roles", "security roles", "responsibility assignment"], "5.2 Information Security Roles and Responsibilities"),
    (["segregation of duties", "separation of duties", "dual control", "four eyes principle"], "5.3 Segregation of Duties"),
    (["management commitment", "management responsibilities", "senior management", "executive sponsor"], "5.4 Management Responsibilities"),
    (["contact with authorities", "law enforcement", "regulatory body", "government contact", "police contact"], "5.5 Contact with Authorities"),
    (["special interest group", "isac", "industry group", "security forum", "peer group"], "5.6 Contact with Special Interest Groups"),
    (["threat intelligence", "threat feed", "cti", "threat data", "ioc", "indicators of compromise"], "5.7 Threat Intelligence"),
    (["project management", "project security", "sdlc governance", "project risk", "security in projects"], "5.8 Information Security in Project Management"),
    (["acceptable use", "aup", "usage policy", "acceptable use policy", "permitted use"], "5.10 Acceptable Use of Information and Other Associated Assets"),
    (["return of assets", "asset return", "exit assets", "equipment return", "device return on exit"], "5.11 Return of Assets"),
    (["data classification", "information classification", "classification scheme", "sensitivity label", "classification policy"], "5.12 Classification of Information"),
    (["data labelling", "information labelling", "label policy", "classification marking", "document marking"], "5.13 Labelling of Information"),
    (["information transfer", "data transfer", "file transfer", "email security", "secure transfer", "sftp protocol", "secure ftp"], "5.14 Information Transfer"),
    (["identity management", "user provisioning", "joiner leaver", "account lifecycle", "iam policy", "user lifecycle"], "5.16 Identity Management"),
    (["access rights", "user permissions", "permission management", "access provisioning", "least privilege access"], "5.18 Access Rights"),
    (["supplier relationship", "vendor relationship", "third party security", "supplier policy", "vendor agreement"], "5.19 Information Security in Supplier Relationships"),
    (["supplier agreement", "vendor contract", "third party contract", "outsourcing agreement", "service contract"], "5.20 Addressing Information Security Within Supplier Agreements"),
    (["ict supply chain", "supply chain security", "hardware supply", "software supply", "supply chain risk"], "5.21 Managing Information Security in The Ict Supply Chain"),
    (["supplier monitoring", "vendor monitoring", "third party review", "supplier audit", "vendor performance"], "5.22 Monitoring, Review and Change Management of Supplier Services"),
    (["cloud service", "cloud security", "saas", "iaas", "paas", "cloud provider", "aws", "azure", "gcp"], "5.23 Information Security for Use of Cloud Services"),
    (["incident assessment", "security event assessment", "incident triage", "event decision", "incident classification"], "5.25 Assessment and Decision on Information Security Events"),
    (["incident response plan", "incident handling", "containment", "eradication", "recovery response"], "5.26 Response to Information Security Incidents"),
    (["lessons learned", "post incident review", "incident review", "learning from incidents", "incident debrief"], "5.27 Learning from Information Security Incidents"),
    (["evidence collection", "forensic", "digital forensic", "chain of custody", "forensic evidence"], "5.28 Collection of Evidence"),
    (["ict continuity", "ict readiness", "it continuity", "system continuity", "technology continuity"], "5.30 Ict Readiness for Business Continuity"),
    (["legal requirement", "statutory requirement", "regulatory requirement", "contractual requirement", "compliance obligation"], "5.31 Legal, Statutory, Regulatory and Contractual Requirements"),
    (["intellectual property", "copyright", "ip rights", "software license", "license management"], "5.32 Intellectual Property Rights"),
    (["independent review", "internal audit", "isms review", "third party audit", "external audit"], "5.35 Independent Review of Information Security"),
    (["compliance check", "policy compliance", "compliance with policy", "compliance review", "standards compliance"], "5.36 Compliance with Policies and Standards for Information Security"),
    (["operating procedures", "sop", "documented procedures", "standard operating procedure", "work instructions"], "5.37 Documented Operating Procedures"),

    # ── Clause 6: People Controls (all previously missing) ────────────────────
    (["employment terms", "terms and conditions", "employment contract", "job agreement", "offer letter"], "6.2 Terms and Conditions of Employment"),
    (["security awareness", "awareness training", "security training", "e-learning", "phishing awareness", "staff training"], "6.3 Information Security Awareness, Education and Training"),
    (["disciplinary", "disciplinary process", "misconduct", "policy violation", "disciplinary action"], "6.4 Disciplinary Process"),
    (["termination", "offboarding", "exit process", "leaver", "resignation", "dismissal", "account deactivation on exit"], "6.5 Responsibilities after Termination or Change of Employment"),
    (["nda", "non-disclosure", "confidentiality agreement", "non disclosure agreement", "confidentiality clause"], "6.6 Confidentiality or Non-disclosure Agreements"),
    (["remote work", "remote working", "work from home", "wfh", "telework", "home working", "vpn policy"], "6.7 Remote Working"),
    (["security event reporting", "report incident", "event reporting", "staff reporting", "how to report"], "6.8 Information Security Event Reporting"),

    # ── Clause 7: Physical Controls (all previously missing) ──────────────────
    (["physical entry", "entry control", "door access", "turnstile", "access control door", "entry point"], "7.2 Physical Entry"),
    (["secure office", "secure room", "secure facility", "server room", "data center access", "office security"], "7.3 Securing Offices, Rooms and Facilities"),
    (["physical monitoring", "cctv", "camera surveillance", "security camera", "video surveillance"], "7.4 Physical Security Monitoring"),
    (["environmental threat", "flood protection", "fire protection", "power protection", "environmental control", "uninterruptible power"], "7.5 Protecting against Physical and Environmental Threats"),
    (["secure area", "working in secure area", "restricted area policy", "secure zone", "clean area"], "7.6 Working in Secure Areas"),
    (["clean desk policy", "clear screen policy", "unattended workstation", "clear desk"], "7.7 Clear Desk and Clear Screen"),
    (["equipment siting", "equipment placement", "server placement", "equipment protection", "rack security"], "7.8 Equipment Siting and Protection"),
    (["off-premises", "assets off-premises", "remote equipment", "equipment offsite", "off site device"], "7.9 Security of Assets Off-premises"),
    (["storage media", "removable media", "usb", "hard drive disposal", "media handling", "removable storage"], "7.10 Storage Media"),
    (["supporting utilities", "power supply failure", "backup generator", "utility failure", "electricity supply", "uninterruptible power supply"], "7.11 Supporting Utilities"),
    (["cabling", "cable security", "network cabling", "structured cabling", "cable management"], "7.12 Cabling Security"),
    (["equipment maintenance", "maintenance schedule", "server maintenance", "hardware maintenance", "preventive maintenance"], "7.13 Equipment Maintenance"),
    (["secure disposal", "equipment disposal", "data destruction", "disk wipe", "degauss", "decommission"], "7.14 Secure Disposal or Re-use of Equipment"),

    # ── Clause 8: Technological Controls (all previously missing) ─────────────
    (["endpoint", "user endpoint", "laptop policy", "mobile device", "mdm", "byod", "endpoint security"], "8.1 User Endpoint Devices"),
    (["information access restriction", "need to know", "access restriction", "data access control"], "8.3 Information Access Restriction"),
    (["source code", "code repository", "git access", "repository access", "source code access", "github", "gitlab"], "8.4 Access to Source Code"),
    (["anti-malware", "edr", "malware protection", "endpoint protection", "av policy"], "8.7 Protection against Malware"),
    (["configuration management", "baseline configuration", "hardening", "cmdb", "config baseline", "secure configuration"], "8.9 Configuration Management"),
    (["data deletion", "information deletion", "secure erase", "data wiping", "data removal", "right to erasure"], "8.10 Information Deletion"),
    (["data masking", "anonymization", "pseudonymization", "masking policy", "data anonymisation"], "8.11 Data Masking"),
    (["dlp", "data leakage prevention", "data loss prevention", "data exfiltration", "information leakage"], "8.12 Data Leakage Prevention"),
    (["system redundancy", "failover mechanism", "high availability setup", "active active", "active passive", "redundant system"], "8.14 Redundancy of Information Processing Facilities"),
    (["log management", "syslog", "audit logs", "event logs", "logging policy", "log collection", "log retention"], "8.15 Logging"),
    (["privileged utility", "admin tools", "utility programs", "system utilities", "privileged software"], "8.18 Use of Privileged Utility Programs"),
    (["software installation", "approved software", "application whitelist", "install policy", "software approval"], "8.19 Installation of Software on Operational Systems"),
    (["network services", "service security", "api gateway", "network service policy", "managed network service"], "8.21 Security of Network Services"),
    (["network segregation", "vlan", "network segmentation", "dmz", "network zone", "micro segmentation"], "8.22 Segregation of Networks"),
    (["web filtering", "url filtering", "proxy", "content filter", "internet filtering", "web proxy"], "8.23 Web Filtering"),
    (["secure development", "sdlc", "secure development lifecycle", "secure development policy"], "8.25 Secure Development Life Cycle"),
    (["application security", "app security requirements", "security requirements", "security in design"], "8.26 Application Security Requirements"),
    (["secure architecture", "security architecture", "engineering principles", "security by design", "architecture review"], "8.27 Secure System Architecture and Engineering Principles"),
    (["secure coding", "code review", "coding standard", "owasp", "sast", "dast", "static analysis"], "8.28 Secure Coding"),
    (["security testing", "acceptance testing", "uat security", "pre-production testing", "security test"], "8.29 Security Testing in Development and Acceptance"),
    (["outsourced development", "third party development", "vendor development", "offshore development"], "8.30 Outsourced Development"),
    (["separation of environments", "dev test prod", "environment separation", "non-production", "staging environment"], "8.31 Separation of Development, Testing and Production Environments"),
    (["change management", "change control", "change request", "change advisory board", "change approval process"], "8.32 Change Management"),
    (["test information", "test data", "test data management", "production data in test", "sanitised test data"], "8.33 Test Information"),
    (["audit testing", "audit protection", "system during audit", "audit tools", "audit environment"], "8.34 Protection of Information Systems During Audit Testing"),
]


def _resolve_control_by_direct_map(text: str, use_cases: List[Dict]) -> Optional[Dict]:
    """Fast O(n) keyword-map lookup before falling back to USE_CASES search.
    Ranks matches by keyword length so specific multi-word phrases (e.g. 'supplier monitoring')
    take precedence over shorter substring matches (e.g. 'monitoring')."""
    text_norm = _normalize(text)
    text_lower = str(text or "").lower()
    matches = []

    for keywords, target_use_case in _DIRECT_KEYWORD_CONTROL_MAP:
        for kw in keywords:
            kw_norm = _normalize(kw)
            if kw_norm and (kw_norm in text_norm or kw.lower() in text_lower):
                matches.append((len(kw_norm), target_use_case))

    if matches:
        # Longest match wins (e.g. 'supplier monitoring' len 19 > 'monitoring' len 10)
        matches.sort(key=lambda x: x[0], reverse=True)
        best_target = matches[0][1]
        target_norm = _normalize(best_target)
        for uc in use_cases:
            if _normalize(uc.get("use_case", "")) == target_norm:
                return uc

    return None



# ── Known ISO 27001 & VAPT control short-IDs (matches 5.15, 8.17, VAPT-1 .. VAPT-15)
_CONTROL_ID_RE = re.compile(
    r'\b(VAPT\s*-?\s*\d{1,2}|\d{1,2}\.\d{1,2}(?:\.\d{1,2})?)\b',
    re.IGNORECASE
)


def _normalize(text: str) -> str:
    """Lowercase and strip punctuation for fuzzy matching."""
    return re.sub(r'[^a-z0-9\s]', ' ', str(text or "").lower()).strip()


def _load_use_cases() -> List[Dict]:
    """Load USE_CASES from controls_data safely."""
    try:
        from src.core.controls_data import USE_CASES
        return USE_CASES
    except ImportError:
        return []


def _resolve_control_by_id(text: str, use_cases: List[Dict]) -> Optional[Dict]:
    """Step 1: Extract control ID directly from text (e.g., '8.16', '5.15', 'VAPT-1', 'ISO 8.17')."""
    if not text:
        return None
    matches = _CONTROL_ID_RE.findall(str(text))
    for m in matches:
        m_norm = re.sub(r'vapt\s*-?\s*', 'VAPT-', m, flags=re.IGNORECASE).strip().upper()
        for uc in use_cases:
            uc_id = str(uc.get("use_case", "")).split(" ")[0].upper()
            if uc_id == m_norm or uc_id == m.upper():
                return uc
    return None




def _resolve_control_by_name(text: str, use_cases: List[Dict]) -> Optional[Dict]:
    """Step 2: Match control by name keywords from USE_CASES labels."""
    norm = _normalize(text)
    if not norm:
        return None

    words = [w for w in norm.split() if len(w) > 2]  # Filter tiny stop words
    if not words:
        return None

    best_uc = None
    best_score = 0

    for uc in use_cases:
        label = _normalize(uc.get("label", "") + " " + uc.get("use_case", ""))
        l_words = set(label.split())
        overlap = len(set(words) & l_words)
        if overlap > best_score:
            best_score = overlap
            best_uc = uc

    # If query is short (1-2 meaningful words like 'Capacity' or 'Screening'), require at least 1 match.
    # Otherwise require at least 2 matches.
    min_required = 1 if len(words) <= 2 else 2
    return best_uc if best_score >= min_required else None


def _resolve_control_by_embedding(text: str, use_cases: List[Dict]) -> Optional[Dict]:
    """Step 3: Cosine similarity against USE_CASES descriptions using cached embeddings."""
    try:
        import numpy as np
        import requests

        # Use Ollama embedding model
        ollama_url = os.environ.get("OLLAMA_HOST", "http://127.0.0.1:11434")
        if not ollama_url.startswith("http"):
            ollama_url = f"http://{ollama_url}"

        embed_model = os.environ.get("EMBED_MODEL", "nomic-embed-text")

        def get_embedding(t: str):
            try:
                resp = requests.post(
                    f"{ollama_url}/api/embeddings",
                    json={"model": embed_model, "prompt": t},
                    timeout=60

                )
                if resp.status_code == 200:
                    return np.array(resp.json().get("embedding", []), dtype=np.float32)
            except Exception:
                pass
            return None

        q_vec = get_embedding(text)
        if q_vec is None or len(q_vec) == 0:
            return None

        best_uc = None
        best_sim = -1.0

        for uc in use_cases:
            desc = f"{uc.get('use_case', '')} {uc.get('label', '')} {uc.get('prompt_hint', '')}"
            uc_vec = get_embedding(desc)
            if uc_vec is None or len(uc_vec) == 0:
                continue
            # Cosine similarity
            denom = (np.linalg.norm(q_vec) * np.linalg.norm(uc_vec))
            if denom == 0:
                continue
            sim = float(np.dot(q_vec, uc_vec) / denom)
            if sim > best_sim:
                best_sim = sim
                best_uc = uc

        return best_uc if best_sim > 0.5 else None

    except Exception as e:
        print(f"[EXCEL SCOPING] Embedding resolution failed: {e}", flush=True)
        return None


def _resolve_control(text: str, use_cases: List[Dict]) -> Dict:
    """
    Resolve ISO control from question/control-name text using 4 fallback steps.
    Returns a dict with control_id, control_label, expected_evidence, prompt_hint, severity.
    Falls back to a generic unknown control if all steps fail.
    """
    uc = (
        _resolve_control_by_id(text, use_cases) or
        _resolve_control_by_direct_map(text, use_cases) or   # Fast keyword map
        _resolve_control_by_name(text, use_cases) or
        _resolve_control_by_embedding(text, use_cases)
    )

    if uc:
        use_case_str = str(uc.get("use_case", ""))
        parts = use_case_str.split(" ", 1)
        ctrl_id = parts[0] if parts else "UNKNOWN"
        return {
            "control_id":        ctrl_id,
            "control_label":     use_case_str,
            "expected_evidence": str(uc.get("expected", "")),
            "prompt_hint":       str(uc.get("prompt_hint", "")),
            "severity":          str(uc.get("severity", "MEDIUM")),
        }

    # Fallback: unknown control
    return {
        "control_id":        "UNKNOWN",
        "control_label":     str(text).strip(),
        "expected_evidence": "",
        "prompt_hint":       f"Evaluate the provided document against: {text}",
        "severity":          "MEDIUM",
    }


def _detect_file_columns(header_row: list) -> List[int]:
    """
    Auto-detect which column indices contain file references based on header keywords.
    Excludes columns that are clearly FILE TYPE / FORMAT columns (e.g. 'File type', 'Format').
    Returns a list of column indices (0-based).
    """
    file_col_indices = []
    for idx, cell in enumerate(header_row):
        cell_norm = _normalize(str(cell) if cell is not None else "")
        has_file_kw = any(kw in cell_norm for kw in _FILE_COL_KEYWORDS)
        has_type_kw = any(kw in cell_norm for kw in _FILE_TYPE_EXCLUSION_KEYWORDS)
        # Include only if it has a file keyword AND does NOT have a type/format keyword
        if has_file_kw and not has_type_kw:
            file_col_indices.append(idx)
    return file_col_indices


def _get_file_column_roles(header_row: list, file_cols: List[int]) -> Dict[int, str]:
    """
    Classifies each detected file column as 'policy', 'evidence', or 'generic' based
    on its header text (e.g. 'Policy (source-grounded)' -> 'policy', 'File name' ->
    'generic'). Lets downstream code tell the LLM which locked file(s) the auditor
    intended as policy proof vs operational evidence proof, instead of locking both
    into one undifferentiated blob and making the LLM re-derive the split blind.
    """
    roles = {}
    for idx in file_cols:
        cell_norm = _normalize(str(header_row[idx]) if idx < len(header_row) else "")
        if "policy" in cell_norm:
            roles[idx] = "policy"
        elif "evidence" in cell_norm:
            roles[idx] = "evidence"
        else:
            roles[idx] = "generic"
    return roles


_ROW_NUMBER_HEADER_LABELS = {"s no", "sno", "sl no", "sl", "no", "sr no", "srno", "#", "row", "row no", "index", ""}


def _detect_question_column(header_row: list) -> int:
    """
    Auto-detect which column holds the audit check question/control name.
    Falls back to column index 1 (second column) if none found.
    """
    question_keywords = {
        "audit", "check", "question", "control", "policy", "observation",
        "item", "objective", "requirement", "whether", "sl", "no"
    }
    for idx, cell in enumerate(header_row):
        cell_norm = _normalize(str(cell) if cell is not None else "")
        # Column 0 is often S.No / row number -- skip ONLY if its header text
        # specifically looks like a row-index label. A real question column that
        # happens to be first (e.g. "Audit check" as literal column 0, which some
        # real checklists use with no separate S.No column) must NOT be skipped --
        # a blanket "always skip column 0" previously caused column detection to
        # fail entirely on that layout, matching nothing and silently defaulting
        # to the wrong column with zero evidence files attached to any control.
        if idx == 0 and cell_norm in _ROW_NUMBER_HEADER_LABELS:
            continue
        for kw in question_keywords:
            if kw in cell_norm:
                return idx
    return 1  # Default: second column


def _detect_supplementary_name_column(header_row: list, question_col: int) -> Optional[int]:
    """
    Some checklists split control identification across two columns, e.g. a short
    "Control" ID column plus a separate "Control Name" column. Detects that second
    column so its text can be combined with the primary question column -- this
    matters when a row has a blank Control ID but a filled Control Name (or vice
    versa): without combining them, a row with the primary column blank looks like
    an empty row and gets silently dropped, even though the other column has
    everything needed to resolve the control.
    """
    for idx, cell in enumerate(header_row):
        if idx == question_col:
            continue
        cell_norm = _normalize(str(cell) if cell is not None else "")
        if "name" in cell_norm and ("control" in cell_norm or "audit" in cell_norm or "check" in cell_norm):
            return idx
    return None


def parse_excel_scoping_checklist(
    file_path: str,
    sheet_name: str = None,
    uploaded_filenames: List[str] = None
) -> List[Dict]:
    """
    Parse an Excel scoping checklist and return a list of checklist items.

    Parameters
    ----------
    file_path : str
        Absolute path to the .xlsx file.
    sheet_name : str, optional
        Sheet to read. If None, tries common names then falls back to first sheet.
    uploaded_filenames : list[str], optional
        List of filenames already uploaded to the audit session.
        Used to fuzzy-match Excel file references to actual uploaded files.

    Returns
    -------
    list[dict]
        Each item:
        {
            "row_index":        int,      # 1-based row number
            "question":         str,      # Audit check question / control name
            "files":            list[str],# Locked filenames for this item
            "control_id":       str,      # e.g. "8.17"
            "control_label":    str,      # e.g. "8.17 Clock Synchronization"
            "expected_evidence":str,
            "prompt_hint":      str,
            "severity":         str,
        }
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel scoping file not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        # ── Determine which sheet to read ─────────────────────────────────────────
        preferred_names = ["Audit Check", "Scoping", "Checklist", "Controls",
                           "Sheet1", "Sheet 1", "Data"]
        target_sheet = None

        if sheet_name and sheet_name in wb.sheetnames:
            target_sheet = wb[sheet_name]
        else:
            for name in preferred_names:
                if name in wb.sheetnames:
                    target_sheet = wb[name]
                    break
            if target_sheet is None:
                target_sheet = wb[wb.sheetnames[0]]

        rows = list(target_sheet.iter_rows(values_only=True))
        if not rows:
            return []

        # ── Detect header row (first non-empty row) ────────────────────────────────
        header_row_idx = 0
        for i, row in enumerate(rows):
            if any(cell is not None and str(cell).strip() for cell in row):
                header_row_idx = i
                break

        header_row = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]

        # ── Auto-detect columns ────────────────────────────────────────────────────
        question_col = _detect_question_column(header_row)
        name_col = _detect_supplementary_name_column(header_row, question_col)
        file_cols = _detect_file_columns(header_row)

        # A column cannot be BOTH the question column and a file column.
        # This happens when the header is just "Policy Name" (matches both sets of keywords).
        # In that case, treat it as question-only (no file columns).
        file_cols = [c for c in file_cols if c != question_col and c != name_col]
        file_col_roles = _get_file_column_roles(header_row, file_cols)

        print(f"[EXCEL PARSER] Sheet: '{target_sheet.title}' | "
              f"Question col: {question_col} ('{header_row[question_col]}') | "
              f"Name col: {name_col} ({header_row[name_col] if name_col is not None else None}) | "
              f"File cols: {file_cols} ({[header_row[c] for c in file_cols]}) | "
              f"Roles: {file_col_roles}",
              flush=True)

        use_cases = _load_use_cases()
        items = []

        # ── Iterate data rows ──────────────────────────────────────────────────────
        for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            # Skip completely empty rows
            if not any(c is not None and str(c).strip() for c in row):
                continue

            # Extract question text -- combine the primary column with the
            # supplementary name column (if any) so a row is still resolvable
            # and non-empty when only one of the two is filled in (e.g. a
            # blank Control ID but a filled Control Name, or vice versa).
            q_cell = row[question_col] if question_col < len(row) else None
            q_text = str(q_cell).strip() if q_cell is not None else ""
            name_text = ""
            if name_col is not None and name_col < len(row) and row[name_col] is not None:
                name_text = str(row[name_col]).strip()
            resolution_text = " ".join(t for t in [q_text, name_text] if t)
            question = name_text or q_text  # prefer the more descriptive text for display

            # Skip header-like rows (e.g., "Audit check" appearing mid-sheet)
            if not resolution_text or _normalize(resolution_text) in {
                "audit check", "question", "control", "sl no", "s.no", "sno",
                "control control name", "control name"
            }:
                continue

            # Extract filenames from all detected file columns, split by the
            # column's role (policy/evidence/generic) so the split can be
            # preserved downstream instead of collapsing into one blob.
            raw_files, raw_policy_files, raw_evidence_files = [], [], []
            for fc in file_cols:
                if fc < len(row) and row[fc] is not None:
                    val = str(row[fc]).strip()
                    if val and val.lower() not in {"file name", "file", "filename", "n/a", "na", "-", ""}:
                        raw_files.append(val)
                        role = file_col_roles.get(fc, "generic")
                        if role == "policy":
                            raw_policy_files.append(val)
                        elif role == "evidence":
                            raw_evidence_files.append(val)

            # Fuzzy-match raw file names to actually uploaded files
            matched_files = _match_filenames(raw_files, uploaded_filenames or [])
            matched_policy_files = _match_filenames(raw_policy_files, uploaded_filenames or [])
            matched_evidence_files = _match_filenames(raw_evidence_files, uploaded_filenames or [])

            # Resolve ISO control from the combined text (ID regex needs the raw
            # "5.15"-style value if present; name-keyword/embedding fallback needs
            # the descriptive text if the ID is blank) -- either source alone may
            # be missing for a given row, so both are passed together.
            ctrl_info = _resolve_control(resolution_text, use_cases)

            items.append({
                "row_index":         row_idx,
                "question":          question,
                "files":             matched_files,
                # Role-split view of the same files, when the sheet has separately
                # named Policy/Evidence columns -- both empty for a generic single
                # "File name" column, since there's no column-level signal to split.
                "policy_files":      matched_policy_files,
                "evidence_files":    matched_evidence_files,
                "raw_file_refs":     raw_files,   # original Excel values (for debugging)
                "control_id":        ctrl_info["control_id"],
                "control_label":     ctrl_info["control_label"],
                "expected_evidence": ctrl_info["expected_evidence"],
                "prompt_hint":       ctrl_info["prompt_hint"],
                "severity":          ctrl_info["severity"],
            })

        print(f"[EXCEL PARSER] Parsed {len(items)} checklist items.", flush=True)
        return items
    finally:
        try:
            wb.close()
        except Exception:
            pass



_FILENAME_EXT_RE = re.compile(r'\.(docx?|pdf|xlsx?|csv|pptx?|txt|png|jpe?g|zip)\b', re.IGNORECASE)
_CITATION_SPLIT_RE = re.compile(r'\s*;\s*|\n')
_CITATION_LEADING_TAG_RE = re.compile(r'^\[[^\]]+\]\s*')


def _looks_like_filename_reference(segment: str) -> bool:
    """
    A segment is only treated as a candidate file citation (matched against
    uploads, or kept as an unresolved reference) if it's filename-shaped -- has
    a recognized extension, or is short enough to plausibly be a citation like
    "Some Policy V17.0 -- p.2." rather than a full prose sentence describing
    what a policy says. Long narrative text with no extension is never kept as
    a fake "locked filename" -- that silently breaks retrieval later, since no
    real file will ever match it.
    """
    if _FILENAME_EXT_RE.search(segment):
        return True
    return len(segment) <= 80


def _split_citation_segments(raw_ref: str) -> List[str]:
    """
    A single cell may cite multiple files at once (e.g. "A.docx; B.pdf -- p.2."),
    or bracket-tag a citation (e.g. "[Published] Some Policy.docx"). Splits on
    ';' / newlines and strips leading bracket tags so each file gets matched
    independently instead of the whole cell being treated as one reference.
    """
    parts = [p.strip() for p in _CITATION_SPLIT_RE.split(raw_ref) if p.strip()]
    cleaned = [_CITATION_LEADING_TAG_RE.sub('', p).strip() for p in parts]
    cleaned = [p for p in cleaned if p]
    return cleaned or [raw_ref]


def _match_filenames(
    raw_refs: List[str],
    uploaded_filenames: List[str]
) -> List[str]:
    """
    Fuzzy-match raw Excel file references to actual uploaded filenames.

    Each raw reference may cite multiple files in one cell, or be pure
    narrative text with no file reference at all -- both are handled by first
    splitting into citation segments (see _split_citation_segments), then
    matching each segment independently.

    Strategy per segment:
    1. Exact match (case-insensitive, extension-stripped)
    2. Partial match -- checks BOTH the full uploaded filename and its
       extension-stripped stem as a substring, since citation text doesn't
       always repeat the file extension (e.g. "...Policy V17.0 -- p.2." cites
       "...Policy V17.0.pdf" without ever writing ".pdf")
    3. If nothing matches, keep the segment only if it's filename-shaped (see
       _looks_like_filename_reference) so the auditor can see an expected file
       that hasn't been uploaded yet. Pure narrative segments are dropped
       instead of being kept as a fake filename reference.
    """
    def _resolve_segment(seg: str) -> Optional[str]:
        seg_norm = _normalize(seg)

        # Step 1: exact match (extension-stripped)
        for upl in uploaded_filenames:
            upl_stem_norm = _normalize(os.path.splitext(upl)[0])
            if seg_norm == upl_stem_norm or seg_norm == _normalize(upl):
                return upl

        # Step 2: partial match -- try both the full filename and its
        # extension-stripped stem, since citations often omit the extension
        best_match, best_len = None, 0
        for upl in uploaded_filenames:
            for candidate in (_normalize(upl), _normalize(os.path.splitext(upl)[0])):
                if not candidate:
                    continue
                if candidate in seg_norm and len(candidate) > best_len:
                    best_len, best_match = len(candidate), upl
                elif seg_norm in candidate and len(seg_norm) > best_len:
                    best_len, best_match = len(seg_norm), upl
        return best_match

    matched = []
    for ref in raw_refs:
        for seg in _split_citation_segments(ref):
            resolved = _resolve_segment(seg) if uploaded_filenames else None
            if resolved:
                matched.append(resolved)
            elif _looks_like_filename_reference(seg):
                # Step 3: keep as an unresolved-but-plausible reference
                matched.append(seg)

    # Deduplicate while preserving order -- the same file can legitimately be
    # cited twice (e.g. once in the Policy column, once in Evidence)
    seen = set()
    result = []
    for f in matched:
        if f not in seen:
            seen.add(f)
            result.append(f)
    return result


def get_locked_filenames_for_control(
    checklist_items: List[Dict],
    control_id: str
) -> List[str]:
    """
    Returns all locked filenames for a given control_id across all checklist items.
    Useful when multiple checklist rows map to the same control (e.g., two NTP rows → 8.17).
    """
    files = []
    for item in checklist_items:
        if item.get("control_id") == control_id:
            files.extend(item.get("files", []))
    # Deduplicate while preserving order
    seen = set()
    result = []
    for f in files:
        if f not in seen:
            seen.add(f)
            result.append(f)
    return result
